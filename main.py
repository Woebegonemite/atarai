# Author Lewis Stuart 201262348

import sys
import time
import random
import numpy as np
import matplotlib.pyplot as plt
from itertools import count
import torch.optim as optim
from torchviz import make_dot
import pandas as pd
from collections import namedtuple
import xlsxwriter
from openpyxl import load_workbook

# Imports all policies that the agent can follow to achieve optimal actions
from policies import *

# Imports strategies for either following the policy or exploring
from strategies import *

# Imports agents that perform the actions based on the policy
from agents import *

# Imports environment handlers that manage interactions with the environment
from environment_handler import *

# Ensures that graph vision can work correctly
import os
os.environ["PATH"] += os.pathsep + "C:\\Program Files\\Graphviz\\bin"

# Ensures that the results will be the same (same starting random seed each time)
random.seed(0)
np.random.seed(0)

# Creates a diagram of the current neural network
show_neural_net = True

# Lists all atari games the agent can interact with
atari_games = ["SpaceInvaders-v0", "CartPole-v0", "Pong-v0", "BreakoutDeterministic-v4"]

# Default game
default_atari_game = "BreakoutDeterministic-v4"

# Stores the optimal parameters for each available atari game
OptimalParameters = namedtuple(
    'OptimalParameters',
    (
     'learning_rate',  # Learning rate of the policy network (how much each change effects the network)
     'epsilon_values',  # Exploration vs Exploration values
     'discount',  # How impactful future rewards are
     'resize',  # The final size of the screen to enter into the neural network
     'crop_values',  # Crop values to shrink down the size of the screen
     'screen_process_type',  # How the environment processes the screen
     'prev_states_queue_size', # How many states to store in the queue for returning the analysed state
     'policy'
     )
)

# Holds all the optimal parameters for all the avaliable Atari games
optimal_game_parameters = {}

# Optimal Pong parameters
optimal_game_parameters["Pong-v0"] = OptimalParameters(
    0.03,
    [1, 0.05, 0.005],
    0.999,
    [80, 50],
    [[0.06, 0.94],[0.17, 0.92]],
    "difference",
    4,
    'DQN'
)

# Optimal Breakout parameters
optimal_game_parameters["BreakoutDeterministic-v4"] = OptimalParameters(
    0.001,
    [1, 0.1, 0.0005],
    0.999,
    [72, 40],
    [[0.05, 0.95], [0.25, 0.95]],
    "append",
    4,
    'DQN_CNN'
)

"""
# Optimal Breakout parameters
optimal_game_parameters["BreakoutDeterministic-v4"] = OptimalParameters(
    0.001,
    [1, 0.1, 0.0025],
    0.999,
    [80, 50],
    [[0.05, 0.95], [0.15, 0.95]],
    "append",
    4,
    'DQN_CNN'
)

"""


# Optimal Breakout parameters
optimal_game_parameters["CartPole-v0"] = OptimalParameters(
    0.001,
    [1, 0.01, 0.01],
    0.999,
    [40, 90],
    [[0, 1], [0.4, 0.8]],
    "difference",
    2,
    'DQN_CNN'
)


# Rendering information
game_FPS = 30
actions_per_second = 15

# Number of states in a batch
batch_size = 256

# Discount value
discount_value = 0.999

# When the weights of the target neural network should be updated with the policy networks weights
# after a set number of episodes
target_update = 10

# The capacity of the replay memory
memory_size = 100000

# Episodes to train
num_training_episodes = 3000

# Updates the plot after so many episodes
plot_update_episode_factor = 20

# How many times to save the current agent progress (saves neural network)
save_target_network_factor = 200

# Render's the agent performing the eps
render_agent = False

# Will set whether to use the user menu
use_menu = False

# Shows the processed images every 100 steps
show_processed_screens = False


# Writes out the essential information of the training episodes to an excel file for each game
def write_results(info_per_episode):

    # Dataframe of the results of each episode
    results_data_frame = pd.DataFrame(info_per_episode)
    results_data_frame.name = "AI results"

    # Converts parameter data into a dataframe
    parameter_data_dict = optimal_game_parameters[default_atari_game]._asdict()
    for key in parameter_data_dict.keys():
        parameter_data_dict[key] = str(parameter_data_dict[key])

    parameter_data_dict["num_training_episodes"] = str(num_training_episodes)

    parameter_data_frame = pd.DataFrame(parameter_data_dict, index=[0])
    parameter_data_frame.name = "Parameter information"

    # Sets Excel file as the atari game
    file_name = default_atari_game + "_results.xlsx"

    # If the file does not exist currently, then it is created
    if not os.path.exists(file_name):
        workbook = xlsxwriter.Workbook(file_name)
        workbook.close()

    # File is opened and dataframes are added to a new sheet
    workbook = load_workbook(file_name)
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        writer.book = workbook
        parameter_data_frame.to_excel(writer, sheet_name='Results_', startrow=1, startcol=0)
        results_data_frame.to_excel(writer, sheet_name='Results_', startrow=parameter_data_frame.shape[0] + 5, startcol=0)
        writer.save()


# Plots the current episode statistics
def plot(info_per_episode, final):

    # Declares the number of steps, total reward and total time for each episode
    steps_per_episode = [episode["num_steps"] for episode in info_per_episode]
    rewards_per_episode = [episode["total_reward"] for episode in info_per_episode]
    total_time = [float(episode["total_time"]) for episode in info_per_episode]
    moving_average = [episode["moving_average"] for episode in info_per_episode]

    # Sets up graph
    plt.figure(2)
    plt.clf()
    plt.title("Reward, steps and total time for each training episode")
    plt.xlabel("Episode number")
    plt.ylabel("Total per element")
    #plt.legend(["Steps", "Reward", "Time (ms)", "Moving average (reward)"])
    plt.locator_params(axis='y', nbins=5)
    plt.locator_params(axis='x', nbins=5)
    point_intervals = round(len(info_per_episode)/10)
    if point_intervals < 1:
        point_intervals = 1
    plt.xticks(np.arange(1, len(info_per_episode), point_intervals))

    # Plots point
    plt.plot(steps_per_episode, '-bx', label="Steps")
    plt.plot(rewards_per_episode, '-rx',  label="Rewards")
    plt.plot(total_time, '-gx',  label="Time")

    # Plots moving averages
    plt.plot(moving_average, '-kx', label=f"Moving average (reward)")
    plt.pause(0.001)

    # Saves the final plot
    if final:
        plt.savefig("Final Analysis")


# Extracts tensors from experiences
def extract_tensors(experiences):

    # Convert batch of Experiences to Experience of batches
    batch = Experience(*zip(*experiences))

    # This has the following format:
    # Experience(state = (1,2,3), action = (1,2,3), next_state = (1,2,3), reward = (1,2,3))

    # Assigns each tensor to a variable
    t1 = torch.cat(batch.state)
    t2 = torch.cat(batch.action)
    t3 = torch.cat(batch.reward)
    t4 = torch.cat(batch.next_state)

    # Returns an tuple of the experiences
    return (t1,t2,t3,t4)


# Trains the agent using deep Q learning
def train_Q_agent(em, agent):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # Establishes the replay memory
    memory = ReplayMemory(memory_size)

    # Screen width and heights are returned
    screen_width = em.get_screen_width()
    screen_height = em.get_screen_height()

    # Uses a deep neural network (without convolution layers)
    if optimal_game_parameters[default_atari_game].policy == "DQN":
        # Sets up input sizes for the networks
        policy_net = DQN(screen_height, screen_width, em.num_actions_available()).to(device)

        # Sets default weights
        policy_net.apply(initialise_weights)

        target_net = DQN(screen_height, screen_width, em.num_actions_available()).to(device)

    # Uses a deep neural network (with convolution layers)
    elif optimal_game_parameters[default_atari_game].policy == "DQN_CNN":
        # Establishes Policy and Target networks
        policy_net = DQN_CNN(screen_height, screen_width, em.num_actions_available()).to(device)

        # Sets default weights
        policy_net.apply(initialise_weights)

        target_net = DQN_CNN(screen_height, screen_width, em.num_actions_available()).to(device)

    else:
        raise Exception("Policy and target networks not established")

    # Sets the weights and biases to be the same for both networks
    target_net.load_state_dict(policy_net.state_dict())

    # Sets the network to not be in training mode (only be used for inference)
    target_net.eval()

    # Sets and optimiser with the values to optimised as the parameters of the policy network, with the learning rate
    optimizer = optim.Adam(params=policy_net.parameters(), lr=optimal_game_parameters[default_atari_game].learning_rate)

    # Stores episode durations
    episode_durations = []

    # Creates a visual representation of the neural network and saves it as 'Policy Network diagram'
    if show_neural_net:
        print("Creating neural network diagram")
        returned_values = policy_net(em.get_state())
        make_dot(returned_values, params=dict(list(policy_net.named_parameters()))).render("Policy Network diagram",
                                                                                           format="png")

    # Iterates over each episode
    for episode in range(num_training_episodes):

        em.reset()

        # Return initial state
        state = em.get_state()

        # Total episode Reward
        episode_reward = 0

        # Start time of the episode
        start = time.time()

        # Iterates through the number of steps in each episode
        for step in count():

            # Returns action
            action = agent.select_action(state, policy_net, episode)

            # Returns reward
            reward = em.take_action(action)

            episode_reward += reward.numpy()[0]

            # Returns next state
            next_state = em.get_state()

            # Adds experience to list of memory
            memory.push(Experience(state, action, next_state, reward))

            # Updates new state
            state = next_state

            # If set, shows how the states are visualised
            if step % 100 == 0 and show_processed_screens:
                next_state_screen = next_state.squeeze(0).permute(1, 2, 0).cpu()

                plt.figure()
                plt.imshow(next_state_screen, interpolation='none')
                plt.title(f'Computer edited screen: {step}')
                plt.show()

                plt.figure()
                plt.imshow(em.render('rgb_array'))
                plt.title(f'Normal standard screen: {step}')
                plt.show()

            # If set, renders the environment on the screen
            if render_agent or episode > 800:
                em.render()

            # Retrieves a sample if possible and assigns this to the variable 'Experiences'
            if memory.can_provide_sample(batch_size):
                experiences = memory.sample(batch_size)

                # Extracts all states, actions, reward and next states into their own tensors
                states, actions, rewards, next_states = extract_tensors(experiences)

                # Extracts the predicted Q values for the states and actions pairs (as predicted by the policy network)
                current_q_values = QValues.get_current(policy_net, states, actions)

                # Extracts next Q values of the best corresponding actions of the target network
                # The target network is used for finding the next best actions
                next_q_values = QValues.get_next(target_net, next_states)

                # Uses formula E[reward + gamma * maxarg(next state)] to update Q values
                target_q_values = (next_q_values * optimal_game_parameters[default_atari_game].discount) + rewards

                # Calculates loss between the current Q values and the target Q values by using the
                # mean squared error as the loss function
                loss = F.mse_loss(current_q_values, target_q_values.unsqueeze(1))

                # Sets the all gradients of all the weights and biases in the policy network to 0
                # As pytorch accumulates gradients every time it is used, it needs to be reset as to not
                # Factor in old gradients and biases
                optimizer.zero_grad()

                # Computes the gradients of loss (error) of all the weights and biases in the policy network
                loss.backward()

                # Updates the weights and biases of the policy network with the gradient computed from the loss
                optimizer.step()

            # Checks if the episode has finished
            if em.done:

                # Statistic data is recorded
                total_time = str(time.time() - start)

                # Finds the 50 episode moving average
                prev_fifty_episodes = episode_durations[-49:]
                prev_rewards = 0
                for prev_episode in prev_fifty_episodes:
                    prev_rewards += prev_episode["total_reward"]

                prev_rewards = round(((prev_rewards + episode_reward)/(len(prev_fifty_episodes)+1)), 2)

                # Includes all episode information (time, reward, steps)
                episode_info = {"num_steps": step, "total_reward": episode_reward,
                                "total_time": total_time[0:total_time.find('.') + 3],
                                "moving_average": prev_rewards}

                # Appends the episode information
                episode_durations.append(episode_info)

                # Prints the current episode information if set
                if use_menu == False:
                    print(f"Current episode: {episode}")
                    print(f"Reward: {episode_reward}")
                    print(f"Steps: {step}")
                    print(f"Moving_average: {prev_rewards}")
                    print(f"Current epsilon: {agent.return_exploration_rate(episode)}")
                    print(f"Time: {total_time[0:total_time.find('.') + 3]}")
                    print()

                # Draws graph depending on the plot update factor
                if episode % plot_update_episode_factor == 0:
                    # Appends the number of steps
                    plot(episode_durations, False)

                # Episode is finished and breaks
                break

        # Checks to see if the target network needs updating by checking if the episode count is
        # a multiple of the target_update value
        if episode % target_update == 0:
            target_net.load_state_dict(policy_net.state_dict())

        if episode % save_target_network_factor == 0:
            torch.save(target_net.state_dict(), f"network_weights/{default_atari_game}_Policy_Network_{episode}")

    # Plots performance before closing
    plot(episode_durations, True)

    # Writes data to excel file
    write_results(episode_durations)

    # Closes environment
    em.close()

    return target_net


# Agent plays against itself
def self_play(policy_net, em, agent):
    # Iterates through a series of steps until the agent either wins or loses
    current_frame = 1

    em.reset()

    # Return initial state
    state = em.get_state()
    while True:
        step_start_time = time.time()

        action = agent.select_exploitative_action(state, policy_net) if current_frame % (game_FPS/actions_per_second) or (game_FPS/actions_per_second) < 2 else torch.tensor([0]).to(device)

        em.take_action(action)

        # Renders environment
        em.render('human')

        # Executes if the agent either wins or loses
        if em.done: break

        # Syncs environment FPS
        computation_time = time.time() - step_start_time
        break_time = 1/game_FPS - computation_time

        if break_time > 0:
            time.sleep(1/game_FPS)

        current_frame += 1


# Lets the agent play the game either solo, against the user or against another agent
def play_game(play_type, policy_net, em, agent):
    em.reset()

    # Single-player agent
    if play_type == 0:
        print("Single player selected")
        self_play(policy_net, em, agent)

    # Agent vs User
    elif play_type == 1:
        pass

    # Agent vs Agent
    elif play_type == 2:
        pass

    else:
        raise ValueError("Type of play must be an int between 0-2")

    # Lets the user restart the game
    while True:
        restart = str(input("Restart game? (Enter Y or N) \n>")).lower().strip()
        if restart == "y" or restart == "yes":
            play_game(play_type, policy_net, em, agent)
            return
        elif restart == "n" or restart == "no":
            return
        else:
            print("Invalid input")
            print()


# Prints all information about the agent
def print_agent_information():
    print()
    print(f"New game: {default_atari_game}")

    # Outputs action and state space
    print(f"Action Space {em.env.action_space}")
    print(f"State Space {em.env.observation_space}")
    print()

    current_game_parameters = optimal_game_parameters[default_atari_game]

    # Parameters for how the agent is trained
    print("Parameters:")
    print(f"Episodes: {num_training_episodes}")
    print(f"Discount factor: {current_game_parameters.discount}")
    print(f"Learning rate: {current_game_parameters.learning_rate}")
    print(f"Policy used: {current_game_parameters.policy}")

    # Parameters for how choices are made
    print("Epsilon values: ")
    print(f"\t-Start: {current_game_parameters.epsilon_values[0]}")
    print(f"\t-Decay: {current_game_parameters.epsilon_values[2]}")
    print(f"\t-End: {current_game_parameters.epsilon_values[1]}")
    print()

    # Screen values for how the agent views the environment
    print("Screen values:")
    print(f"\t-Crop width percentage: {current_game_parameters.crop_values[0]}")
    print(f"\t-Crop height percentage: {current_game_parameters.crop_values[1]}")
    print(f"Screen resize: {current_game_parameters.resize}")
    print()

    # State processing types are displayed
    print("State processing types:")
    print(f"Number of state queue: {current_game_parameters.prev_states_queue_size}")
    print(f"States analysis type: '{current_game_parameters.screen_process_type}'")
    print()

    # CUDA Information is displayed
    print(f"{'GPU' if torch.cuda.is_available() else 'CPU'} used as primary training device")
    print(f"Torch version: {torch.version}")
    if torch.cuda.is_available():
        print(f"Torch Cuda version: {torch.version.cuda}")
    print()

    print("Running Agent")


if __name__ == '__main__':
    arguements = sys.argv

    # If the user menu is set, passed parameters to the program are evaluated
    if use_menu:
        # If 5 arguments are not passed, then the default arguments are passed
        if len(arguements) != 5:
            print("No arguments: default settings being applied\n")
            play_type = 0
            train = True
            while True:
                reset = str(input("Would you like to reset agent's learning?\n>")).lower().strip()
                if reset == "y" or reset == "yes":
                    reset_agent = True
                    break
                elif reset == "n" or reset == "no":
                    reset_agent = False
                    break
                else:
                    print("Invalid input")
                    print()

        # Otherwise the game settings are set to the passed arguments
        else:
            default_atari_game = sys.argv[1].lower().strip()
            play_type = sys.argv[2]
            try:
                train = bool(sys.argv[3])
                reset_agent = bool(sys.argv[4])
            except BaseException:
                raise TypeError("Train and reset value must be either True or False")

        if play_type not in [0, 1, 2] or type(play_type) != int:
            raise ValueError("Type of play must be an int between 0-2")
        elif default_atari_game not in atari_games:
            raise ValueError(f"Passed Atari game '{default_atari_game}' could not be found")

    # Used if debugging and just training
    else:
        # Sets default settings
        play_type = 0
        train = True

    try:
        # The percentages to crop the screen for returning a state
        crop_width = optimal_game_parameters[default_atari_game].crop_values[0]
        crop_height = optimal_game_parameters[default_atari_game].crop_values[1]

        # Resizes the image for output
        resize = optimal_game_parameters[default_atari_game].resize

        screen_process_type = optimal_game_parameters[default_atari_game].screen_process_type

        prev_states_queue_size = optimal_game_parameters[default_atari_game].prev_states_queue_size

        # Environment is set to the passed atari game
        em = EnvironmentManager(default_atari_game, [crop_width, crop_height], resize, screen_process_type, prev_states_queue_size)

        # Action strategy is set
        episilon_values = optimal_game_parameters[default_atari_game].epsilon_values
        strategy = EpsilonGreedyStrategy(episilon_values[0], episilon_values[1], episilon_values[2])

        # Agent is created
        agent = Agent(strategy, em.num_actions_available())

        print(f"All atari games: {em.return_avaliable_atari_games()}")

    except BaseException:
        raise Exception("Failed to load gym environment and agent")

    print_agent_information()

    # Trains the agent if the user has selected to do so
    if train:
        # Total time that the agent has been running
        start_run_time = time.time()
        policy_net = train_Q_agent(em, agent)
        final_run_time = str(time.time() - start_run_time)
        torch.save(policy_net.state_dict(), f"network_weights/{default_atari_game}_Policy_Network_Final")

        pause = input("Agent has finished, enter to continue to normal play \n>")
        print()

    # Attempts to load in a previous deep Q network
    else:
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        policy_net = DQN(em.get_screen_height(), em.get_screen_width(), em.num_actions_available()).to(device)

        # Attempts to load the specific game DQN
        storage = torch.load(f"network_weights/{default_atari_game}_Policy_Network_Final")
        if not storage:
            raise FileNotFoundError(f"Could not load in neural network for game {default_atari_game}, please restart and train a new one")
        policy_net.load_state_dict(storage)

    # Agent plays the game according to the user play type input
    play_game(play_type, policy_net, em, agent)
